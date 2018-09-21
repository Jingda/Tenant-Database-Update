
# coding: utf-8

# In[ ]:


import time
## start = time.time()

from datetime import datetime
start = datetime.now()

import os
import re
import math
import numpy as np
import pandas as pd
from difflib import SequenceMatcher

from copy import copy
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


# In[ ]:


## better use this kind of link for shared disk
path = '\\\\EgnyteDrive\\projectphx\\Shared\\ACQUISITIONS DEPT\\Asset Management\\Tenant Database\\'
path_info = path + 'info from Yardi\\'
files = [ f for f in os.listdir(path_info) if os.path.isfile( os.path.join(path_info, f) ) ]


# In[ ]:


## to enter new password to Yardi if necessary
wb_access = load_workbook(path_info + 'access.xlsx')
sheet_access = wb_access['Sheet1']

wish = input('The username is bjohnson. Do you want to update the password of this account of Yardi? ')
wish = wish.lower()
str_wish = ['y', 'yes', 'n', 'no']

while wish not in str_wish:
    wish = input("Please enter 'yes' or 'no'.\nDo you want to update the password to Yardi? ")

if (wish == 'yes') | (wish == 'y'):
    password = input('Please enter your password: ')
    print( 'The password you entered is {}.'.format(password) )
    ## print( 'The password you entered is {}.'.format( '*******' + password[-4:] ) )
    sheet_access.cell(row = 1, column = 1).value = password
else:
    password = sheet_access.cell(row = 1, column = 1).value

wb_access.save(path_info + 'access.xlsx')


# In[ ]:


## to find the latest downloaded information from Yardi
print('\nScanning Information From Yardi ...')

##
ar = 365
for f in files:
    if 'AgingSummary' in f:
        print(f)
        date = re.search( r'(AgingSummary)_(.+)(.xls).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days < ar:
            ar = diff.days
## print(ar)
if ar < 1:
    ar = 0

##
rent_roll = 365
for f in files:
    if 'RentRoll' in f:
        print(f)
        date = re.search( r'(RentRoll)(.+).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days < rent_roll:
            rent_roll = diff.days
## print(rent_roll)
if rent_roll < 1:
    rent_roll = 0

##
tenancy_schedule = 365
for f in files:
    if 'TenancyScheduleII' in f:
        print(f)
        date = re.search( r'(TenancyScheduleII)(.+).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days < tenancy_schedule:
            tenancy_schedule = diff.days
## print(tenancy_schedule)
if tenancy_schedule < 1:
    tenancy_schedule = 0

##
unit_vacancy = 365
lapse = datetime.max - datetime.min
for f in files:
    if 'UnitVacancy' in f:
        print(f)
        date = re.search( r'(UnitVacancy)(.+).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days < unit_vacancy:
            unit_vacancy = diff.days
## print(unit_vacancy)
if unit_vacancy < 1:
    unit_vacancy = 0
    
print('Finished. \n')


# In[ ]:


## to download informatoin from Yardi
profile = webdriver.FirefoxProfile()
profile.set_preference('browser.download.panel.shown', False)
## default to /home/user/Downloads, profile.set_Preference('browser.download.folderList', 1)
profile.set_preference('browser.download.folderList', 2)
profile.set_preference('browser.download.dir', path_info)
profile.set_preference('browser.helperApps.neverAsk.openFile', 'text/csv, application/x-msexcel, application/excel, application/x-excel, application/vnd.ms-excel, image/png, image/jpeg, text/html, text/plain, application/msword, application/xml')
profile.set_preference('browser.helperApps.neverAsk.saveToDisk', 'text/csv, application/x-msexcel, application/excel, application/x-excel, application/vnd.ms-excel, image/png, mage/jpeg, text/html, text/plain, application/msword, application/xml')
profile.set_preference('browser.download.manager.alertOnEXEOpen', False)
profile.set_preference('browser.download.manager.focusWhenStarting', False)
profile.set_preference('browser.download.manager.useWindow', False)
profile.set_preference('browser.download.manager.showAlertOnComplete', False)
profile.set_preference('browser.download.manager.closeWhenDone', False)

## capabilities = webdriver.DesiredCapabilities().FIREFOX
## capabilities['marionette'] = True
## binary = FirefoxBinary(r'\Mozilla Firefox\firefox.exe')
## driver = webdriver.Firefox(firefox_profile = profile, firefox_binary = binary, capabilities = capabilities)
driver = webdriver.Firefox(firefox_profile = profile)
## driver = webdriver.Chrome()
## driver.set_page_load_timeout(300)
url_login = r'https://www.yardiasptx10.com/47709reit/pages/LoginAdvanced.aspx'
driver.get(url_login)

driver.find_element_by_id('Username').send_keys('bjohnson')

## original codes to block display style
## field = driver.find_element_by_id("selectedFile")
## driver.execute_script("arguments[0].style.display = 'block';", field)
## field = driver.find_element_by_id("selectedFile")
## field.send_keys('/Users/knightfox/Desktop/file.txt')

psw = driver.find_element_by_id('Password')
driver.execute_script( "arguments[0].style.display = 'block';", psw )
psw = driver.find_element_by_id('Password')
psw.send_keys(password)

driver.find_element_by_id('cmdLogin1').click()
## psw.submit()

## incomplete access
## analytics = driver.find_element_by_id("mi3")
## analytics.click()
## financial = driver.find_element_by_id("mi3-1")
## financial_hover = ActionChains(driver).move_to_element(financial)
## financial_hover.perform()

## download rent roll
url_financial_analytics = r'https://www.yardiasptx10.com/47709reit/pages/CommReportPropertySummary.aspx'
driver.get(url_financial_analytics)

## it is very important to print(driver.page_source) to show id is PropertyId_LookupCode instead of PropertyID_LookupCode 
## from inspection
## not working, form = WebDriverWait(driver, 10).until( EC.presence_of_element_located( (By.ID, 'Form1') ) )
## not working, driver.switch_to.frame(0)
## print(driver.page_source)
property_id = driver.find_element_by_id('PropertyId_LookupCode')
driver.execute_script( "arguments[0].value = ''", property_id )
ids_property_str = '310^315^340^375^420^450^455^505^515^520^521^522^523^526^527^528^529^530^531^532^533^534^535^536^537^538^539^540^541^543^544^545^546^547^549^550^551^552^553^554^555^556^557^558^559^560^561^562^563^564^565^566^567^1300^1320^1325^1355^1380^1385^1390^1410^1465^1480^1510^1524^1525^1548'
property_id.send_keys(ids_property_str)

from_date = driver.find_element_by_id('FromDate_TextBox')
driver.execute_script( "arguments[0].value = ''", from_date )
from_date.send_keys( datetime.now().strftime('%m/%d/%Y') )

show_detail = driver.find_element_by_id('chkIsDetail_CheckBox')
if not show_detail.is_selected():
    show_detail.click()

if rent_roll != 0:
    print('Downloading Rent Roll ...')
    driver.find_element_by_id('Excel_Button').click()

    ## it is the only way to wait for several seconds without connection abort due to anti-virus software 
    for i in range(15):
        time.sleep(4)
        driver.find_element_by_id('Body1')
        
    print('Finished.')
    
## download unit vacancy
report_type = driver.find_element_by_id('ReportType_DropDownList')
for option in report_type.find_elements_by_tag_name('option'):
    if option.text == 'Unit Vacancy':
        option.click()
        break

if unit_vacancy != 0:
    print('Downloading Unit Vacancy ...')
    driver.find_element_by_id('Excel_Button').click()

    for i in range(5):
        time.sleep(4)
        driver.find_element_by_id('Body1')
    
    print('Finished.')

## download aging summary
url_ar_analytics = r'https://www.yardiasptx10.com/47709reit/pages/ArAnalytics.aspx?Programtype=2'
driver.get(url_ar_analytics)

property_id = driver.find_element_by_id('PropertyLookup_LookupCode')
driver.execute_script( "arguments[0].value = ''", property_id )
property_id.send_keys(ids_property_str)

report_type = driver.find_element_by_id('cmbReportType_DropDownList')
for option in report_type.find_elements_by_tag_name('option'):
    if option.text == 'Aging Summary':
        option.click()
        break
        
group_type = driver.find_element_by_id('cmbGroupby_DropDownList')
for option in group_type.find_elements_by_tag_name('option'):
    if option.text == 'Tenant':
        option.click()
        break

## The initial statuses of options of all three 'Current', 'Past', 'Future' are selected.
report_type = driver.find_element_by_id('tenstatus_MultiSelect')
for option in report_type.find_elements_by_tag_name('option'):
    if option.text == 'Current':
        ActionChains(driver).double_click(option).perform()
        break
select = Select(report_type)
## select.select_by_visible_text( 'Current' )
select.select_by_visible_text( 'Future' )

## to select only 'Current'
## report_type = driver.find_element_by_id('tenstatus_MultiSelect')
## for option in report_type.find_elements_by_tag_name('option'):
    ## if option.text == 'Current':
        ## ActionChains(driver).double_click(option).perform()
        ## break
        
run_for_today = driver.find_element_by_id('ysiToday_CheckBox')
if not run_for_today.is_selected():
    run_for_today.click()

if ar != 0:
    print('Downloading AR ...')
    driver.find_element_by_id('btnExcel_Button').click()

    for i in range(15):
        time.sleep(4)
        driver.find_element_by_id('Body1')
    
    print('Finished.')
    
## download DBA and sales group
url_tenancy_schedule = r'https://www.yardiasptx10.com/47709reit/pages/CommTenancyScheduleSummaryF2.aspx'
driver.get(url_tenancy_schedule)

property_id = driver.find_element_by_id('PropertyId_LookupCode')
driver.execute_script( "arguments[0].value = ''", property_id )
property_id.send_keys(ids_property_str)

sale_group = driver.find_element_by_id('SalesGroup_LookupCode')
sale_groups = 'beauty^dining^educate^enter^grocery^health^other^service^special'
driver.execute_script( "arguments[0].value = ''", sale_group )
sale_group.send_keys(sale_groups)

from_date = driver.find_element_by_id('FromDate_TextBox')
driver.execute_script( "arguments[0].value = ''", from_date )
from_date.send_keys( datetime.now().strftime('%m/%d/%Y') )

future_lease = driver.find_element_by_id('ShowFutureActiveLease_DropDownList')
for option in future_lease.find_elements_by_tag_name('option'):
    if option.text == 'No':
        option.click()
        break
        
pending_amendments = driver.find_element_by_id('ShowPendingAmendments_DropDownList')
for option in pending_amendments.find_elements_by_tag_name('option'):
    if option.text == 'No':
        option.click()
        break

dba = driver.find_element_by_id('DBA_CheckBox')
if not dba.is_selected():
    dba.click()

pending_amendments = driver.find_element_by_id('CustomTables_DropDownList')
for option in pending_amendments.find_elements_by_tag_name('option'):
    if option.text == 'Sales Group':
        option.click()
        break

lease_type = driver.find_element_by_id('LeaseTypeSummary_CheckBox')
if lease_type.is_selected():
    lease_type.click()
    
charge_code = driver.find_element_by_id('ChargeCodeSummary_CheckBox')
if charge_code.is_selected():
    charge_code.click()
    
occupancy = driver.find_element_by_id('Occupancy_CheckBox')
if occupancy.is_selected():
    occupancy.click()

unit_type = driver.find_element_by_id('chkUnitType_CheckBox')
if unit_type.is_selected():
    unit_type.click()

spaces = driver.find_element_by_id('Spaces_CheckBox')
if spaces.is_selected():
    spaces.click()
    
customer = driver.find_element_by_id('Customer_CheckBox')
if customer.is_selected():
    customer.click()
    
rent_schedule = driver.find_element_by_id('RentSchedule_CheckBox')
if rent_schedule.is_selected():
    rent_schedule.click()

recovery_schedule = driver.find_element_by_id('RecoverySchedule_CheckBox')
if recovery_schedule.is_selected():
    recovery_schedule.click()

retail_schedule = driver.find_element_by_id('RetailSchedule_CheckBox')
if retail_schedule.is_selected():
    retail_schedule.click()
    
retail_MAT = driver.find_element_by_id('Retail_MAT_CheckBox')
if retail_MAT.is_selected():
    retail_MAT.click()

amendments = driver.find_element_by_id('Amendments_CheckBox')
if amendments.is_selected():
    amendments.click()

proposal = driver.find_element_by_id('Proposal_CheckBox')
if proposal.is_selected():
    proposal.click()

market_rent = driver.find_element_by_id('MarketRent_CheckBox')
if market_rent.is_selected():
    market_rent.click()
    
if tenancy_schedule != 0:
    print('Downloading Tenancy Schedule II ...')
    driver.find_element_by_id('Excel_Button').click()

    for i in range(15):
        time.sleep(4)
        driver.find_element_by_id('Body1')
        
    print('Finished. \n')

## to close the Firefox browser window
driver.quit()


# In[ ]:


## to find the latest Tenant Database files
files = [ f for f in os.listdir(path) 
          if os.path.isfile( os.path.join(path, f) ) ]

files_info = [ f for f in os.listdir(path_info) 
          if os.path.isfile( os.path.join(path_info, f) ) ]

print('Scanning Existing Tenant Database Files... ')

##
lapse = datetime.max - datetime.min
lapse_last = datetime.max - datetime.min
for f in files:
    if 'Tenant Database' in f:
        print(f)
        date = re.search( r'(Tenant Database) (.+).(xlsx)', f ).groups()[1]
        try:
            diff = datetime.now() - datetime.strptime(date, '%m-%d-%Y')
            ## print(diff)
            if diff.days > 60:
                os.remove( os.path.join(path, f) )
            if diff < lapse:
                lapse = diff
                path_tenant_database_origin = os.path.join(path, f)
            if (diff < lapse_last) & (diff.days > 1):
                lapse_last = diff
                path_tenant_database_last = os.path.join(path, f)
                date_last = date
        except:
            pass
## print(path_tenant_database_origin)
## print(path_tenant_database_last)

##
lapse = datetime.max - datetime.min
for f in files_info:
    if 'AgingSummary' in f:
        ## print(f)
        date = re.search( r'(AgingSummary)_(.+)(.xls).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days > 7:
            os.remove( os.path.join(path_info, f) )
        if diff < lapse:
            lapse = diff
            path_aging_summary = os.path.join(path_info, f)
## print(path_aging_summary)

##
lapse = datetime.max - datetime.min
for f in files_info:
    if 'RentRoll' in f:
        ## print(f)
        date = re.search( r'(RentRoll)(.+).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days > 7:
            os.remove( os.path.join(path_info, f) )
        if diff < lapse:
            lapse = diff
            path_rent_roll = os.path.join(path_info, f)
## print(path_rent_roll)

##
lapse = datetime.max - datetime.min
for f in files_info:
    if 'TenancyScheduleII' in f:
        ## print(f)
        date = re.search( r'(TenancyScheduleII)(.+).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days > 7:
            os.remove( os.path.join(path_info, f) )
        if diff < lapse:
            lapse = diff
            path_tenancy_schedule = os.path.join(path_info, f)
## print(path_tenancy_schedule)

##
lapse = datetime.max - datetime.min
for f in files_info:
    if 'UnitVacancy' in f:
        ## print(f)
        date = re.search( r'(UnitVacancy)(.+).(xlsx)', f ).groups()[1]
        diff = datetime.now() - datetime.strptime(date, '%m_%d_%Y')
        ## print(diff)
        if diff.days > 7:
            os.remove( os.path.join(path_info, f) )
        if diff < lapse:
            lapse = diff
            path_unit_vacancy = os.path.join(path_info, f)
## print(path_unit_vacancy)

print('Finished. \n')


# In[ ]:


## to update Rent Roll
print('Updating Rent Roll ...')

ids_property = [310, 315, 340, 375, 420, 450, 455, 505, 515, 520, 521, 522, 523, 526, 527, 528, 529, 530, 531, 532, 533, 534, 
                535, 536, 537, 538, 539, 540, 541, 543, 544, 545, 546, 547, 549, 550, 551, 552, 553, 554, 555, 556, 557, 558, 
                559, 560, 561, 562, 563, 564, 565, 566, 567, 1300, 1320, 1325, 1355, 1380, 1385, 1390, 1410, 1465, 1480, 1510, 
                1524, 1525, 1548]

dic_region = {'Houston': 'HOU', 'Spring': 'HOU', 'Pasadena': 'HOU', 'Sugar Land': 'HOU',
              'Dallas': 'DFW', 'Frisco': 'DFW', 'Plano': 'DFW', 'Fort Worth': 'DFW', 'Keller': 'DFW', 'McKinney': 'DFW', 
              'Phoenix': 'PHX', 'Carefree': 'PHX', 'Buffalo Grove': 'PHX', 'Scottsdale': 'PHX', 'Gilbert': 'PHX', 
              'Mesa': 'PHX', 'Chandler': 'PHX', 'Anthem': 'PHX', 'Fountain Hills': 'PHX', 
              'San Antonio': 'CTX', 'Austin': 'CTX'}

## file = r'Z:\Shared\ACQUISITIONS DEPT\Asset Management\Tenant Database\RentRoll.xlsx'
## Or use raw string, file = 'C:\\Users\\acquisitionsdata\\Desktop\\Rent Roll\\Rent Roll.xlsx'
df = pd.read_excel(path_rent_roll, sheet_name = 'Report1')
df.columns = df.iloc[1]

## to get rid of rows without tenants information
idx_pure = [ idx for idx, property in enumerate( df[' Property '] ) if re.match( '\d+', str(property) ) ]
df = df.loc[idx_pure]

## to separate 
dic_property = {}
for idx, x in enumerate( df[' Property '] ):
    if not x.isdigit():
        key = re.search( r'(\d+ - .+) - (\d+),(\D+)', x ).groups()[1]
        value = ( re.search( r'(\d+ - .+) - (\d+),(\D+)', x ).groups()[0], re.search( r'(\d+ - .+) - (\d+),(\D+)', x ).groups()[2] )
        dic_property[key] = value

## to get ride of rows containing only property title
df = df[ df[' Property '].apply( lambda x: x.isdigit() ) ]

df.insert(loc = 0, column = ' Count ', value = 1)

df.insert( loc = df.columns.get_loc(' Count ') + 1, column = ' WSR/PS ', value = '' )
df[' WSR/PS '] = df[' Property '].apply( lambda x: np.where( int(x) < 1000, 'WS', 'PS' ) )

##
df.insert( loc = df.columns.get_loc(' Property ') + 1, column = ' Region ', value = '' )
df[' Region '] = df[' Property '].apply( lambda x: dic_region[ dic_property[x][1] ] )

df.insert( loc = df.columns.get_loc(' Lease '), column = ' DBA Name ', value = '' )

df.insert( loc = df.columns.get_loc(' Lease ') + 1, column = ' Tenant Classification ', value = '' )

df.insert( loc = df.columns.get_loc(' Tenant Classification ') + 1, column = ' Legacy/Acquired/New ', value = 'Legacy' )

df.insert( loc = df.columns.get_loc(' Legacy/Acquired/New ') + 1, column = ' Public/Private ', value = 'Private' )
## df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Public/Private ' ] = ''

df.insert( loc = df.columns.get_loc(' Public/Private ') + 1, column = ' Franchise (Y/N) ', value = 'N' )
## df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Franchise (Y/N) ' ] = ''

df.insert( loc = df.columns.get_loc(' Franchise (Y/N) ') + 1, column = ' Single Tenant Building ', value = 'N' )
## df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Single Tenant Building ' ] = ''

##
df.insert( loc = df.columns.get_loc(' Area ') + 1, column = ' Suite Size Classification ', value = '' )
## conditions = [ df[' Area '] < 3000, 
             ## ( df[' Area '] >= 3000 ) & ( df[' Area '] <= 5000 ), 
             ## ( df[' Area '] > 5000 ) & ( df[' Area '] <= 10000 ), 
             ## ( df[' Area '] > 10000 ) & ( df[' Area '] <= 20000 ), 
             ## ( df[' Area '] > 20000 ) ]
## choices = ['Under 3k', '3k - 5k', '5k - 10k', '10k - 20k', 'Over 20k']
## df[' Suite Size Classification '] = np.select(conditions, choices, default = '')

##
df.loc[ df[' Lease From '].isnull(), ' Lease From ' ] = datetime.min
df.loc[ df[' Lease To '].isnull(), ' Lease To ' ] = datetime.max
## ( datetime.now() - df[' Lease From '] ).apply( lambda x: x.days <= 180 )
## ( df[' Lease To '] - datetime.now() ).apply( lambda x: x.days <= 180 )
## df[' Lease To '] = df[' Lease To '].apply( lambda x: np.where( str(x) > '2800', 'NaT', x ) )
## df[' Lease From '] = pd.to_datetime( df[' Lease From '], format = '%Y-%m-%d' )
## df[' Lease To '] = pd.to_datetime( df[' Lease To '], format = '%Y-%m-%d' )

num_row = df.shape[0]
df = df.reset_index(drop = True)

print('Finished.')


# In[ ]:


## df
## dic_property


# In[ ]:


## to update AR
print('Updating AR ...')

df_AR = pd.read_excel(path_aging_summary, sheet_name = 'Report1')
df_AR.columns = df_AR.iloc[1]
df_AR = df_AR[3:]
df_AR = df_AR[ df_AR['Status'].notnull() ]
df_AR['Property'] = df_AR['Property'].apply( lambda x: re.search( r'(.+ - .+) - (\d+)', x ).groups()[1] )
df_AR.drop('Lease', axis = 1, inplace = True)
df_AR = df_AR.rename( columns = {'Property': ' Property '} )
df_AR = df_AR.rename( columns = {'Name': ' Lease '} )

df_AR_future = df_AR[ df_AR['Status'] == 'Future' ]
df_AR_future = df_AR_future.reset_index(drop = True)
df_AR_future[' Lease '] = df_AR_future[' Lease '].apply(lambda x: ' - Future - ' + x )

df_AR_current = df_AR[ df_AR['Status'] == 'Current' ]
df_AR_current = df_AR_current.reset_index(drop = True)


# In[ ]:


## df_AR_future


# In[ ]:


df = pd.merge( left = df, right = df_AR_current[ [' Property ', ' Lease ', 'Current'] ], on = [' Property ', ' Lease '], how = 'left')
df.drop_duplicates( subset = [' Property ', ' Unit(s) ', ' Lease '], keep = 'last', inplace = True )
df = df.reset_index(drop = True)

print('Finished.')


# In[ ]:


## df
## id_future


# In[ ]:


## to update information from the latest Tenant Database file
df_origin = pd.read_excel(path_tenant_database_origin, sheet_name = 'TENANT')
df_origin.columns = df_origin.iloc[9]
df_origin = df_origin[10:-1]
## df_origin = df_origin[ df_origin['Region'].notnull() ]
df_origin[' Property '] = df_origin[' Property '].apply( lambda x: re.search( r'(\d+) - (.+)', x ).groups()[0] )
df_origin = df_origin.rename( columns = {'Lease Name': ' Lease '} )
df_origin[' Lease '] = df_origin[' Lease '].apply(lambda x: 'VACANT' if 'VACANT' in x else x)
df_origin = df_origin.reset_index(drop = True)

##
## df_origin1 = df_origin.iloc[ :, [ df_origin.columns.get_loc(' Property '), df_origin.columns.get_loc(' Unit(s) '), df_origin.columns.get_loc(' Lease ') ] ]
df_origin10 = df_origin.iloc[ :, [ df_origin.columns.get_loc(' Property '), df_origin.columns.get_loc(' Lease ') ] ]
df_origin20 = df_origin.iloc[ :, [ df_origin.columns.get_loc(' Property '), df_origin.columns.get_loc(' Unit(s) ') ]]

df_origin1 = df_origin.iloc[ :, df_origin.columns.get_loc('Quick Service / Sit Down'):-1 ]
df_origin_select1 = pd.concat( [df_origin10, df_origin1], axis = 1 )

##
df_origin2 = df_origin.iloc[ :, df_origin.columns.get_loc('Legacy / Acquired / New'):df_origin.columns.get_loc(' Lease Type ') ]
df_origin_select2 = pd.concat( [df_origin10, df_origin2], axis = 1 )

##
df_origin3 = df_origin.loc[:, 'Suite Size Classification']
df_origin_select3 = pd.concat( [df_origin20, df_origin3], axis = 1 )


# In[ ]:


## df_origin.iloc
## df_origin.index[ df_origin['Suite Size Classification'] == 'Plug' ]


# In[ ]:


## 
## df = pd.merge( left = df, right = df_origin_select1, on = [' Property ', ' Unit(s) ', ' Lease '], how = 'left' )
df = pd.merge( left = df, right = df_origin_select1, on = [' Property ', ' Lease '], how = 'left' )
df.drop_duplicates( subset = [' Property ', ' Unit(s) ', ' Lease '], keep = 'last', inplace = True )
## df = pd.merge( left = df, right = df_origin_select2, on = [' Property ', ' Unit(s) ', ' Lease '], how = 'left' )
df = pd.merge( left = df, right = df_origin_select2, on = [' Property ', ' Lease '], how = 'left' )
df.drop_duplicates( subset = [' Property ', ' Unit(s) ', ' Lease '], keep = 'last', inplace = True )
## df = pd.merge( left = df, right = df_origin_select3, on = [' Property ', ' Unit(s) ', ' Lease '], how = 'left' )
df = pd.merge( left = df, right = df_origin_select3, on = [' Property ', ' Unit(s) '], how = 'left' )
df.drop_duplicates( subset = [' Property ', ' Unit(s) ', ' Lease '], keep = 'last', inplace = True )
df = df.reset_index(drop = True)

df[' Legacy/Acquired/New '] = df['Legacy / Acquired / New']
df.drop('Legacy / Acquired / New', axis = 1, inplace = True)
df[' Public/Private '] = df['Public / Private']
df.drop('Public / Private', axis = 1, inplace = True)
df[' Franchise (Y/N) '] = df['Franchise (Y/N)']
df.drop('Franchise (Y/N)', axis = 1, inplace = True)
df[' Single Tenant Building '] = df['Single Tenant Building']
df.drop('Single Tenant Building', axis = 1, inplace = True)
df[' Suite Size Classification '] = df['Suite Size Classification']
df.drop('Suite Size Classification', axis = 1, inplace = True)

df[' Suite Size Classification '] = df[' Suite Size Classification '].astype(str)
for row in range(num_row):
    s = df.loc[row, ' Suite Size Classification ']
    if (not s) | (s == 'nan'):
        if df.loc[row, ' Area '] < 3000:
            df.loc[row, ' Suite Size Classification '] = 'Under 3k'
        elif ( df.loc[row, ' Area '] >= 3000 ) & ( df.loc[row, ' Area '] <= 5000 ):
            df.loc[row, ' Suite Size Classification '] = '3k - 5k'
        elif ( df.loc[row, ' Area '] > 5000 ) & ( df.loc[row, ' Area '] <= 10000 ):
            df.loc[row, ' Suite Size Classification '] = '5k - 10k'
        elif ( df.loc[row, ' Area '] > 10000 ) & ( df.loc[row, ' Area '] <= 20000 ):
            df.loc[row, ' Suite Size Classification '] = '10k - 20k'
        else:
            df.loc[row, ' Suite Size Classification '] = 'Over 20k'
    elif ( s[0].isdigit() ) | ( s[0] == 'U' ) | ( s[0] == 'O' ):
        if df.loc[row, ' Area '] < 3000:
            df.loc[row, ' Suite Size Classification '] = 'Under 3k'
        elif ( df.loc[row, ' Area '] >= 3000 ) & ( df.loc[row, ' Area '] <= 5000 ):
            df.loc[row, ' Suite Size Classification '] = '3k - 5k'
        elif ( df.loc[row, ' Area '] > 5000 ) & ( df.loc[row, ' Area '] <= 10000 ):
            df.loc[row, ' Suite Size Classification '] = '5k - 10k'
        elif ( df.loc[row, ' Area '] > 10000 ) & ( df.loc[row, ' Area '] <= 20000 ):
            df.loc[row, ' Suite Size Classification '] = '10k - 20k'
        else:
            df.loc[row, ' Suite Size Classification '] = 'Over 20k'


# In[ ]:


## df


# In[ ]:


## to update Tenancy Schedule II
print('Updating Tenancy Schedule II ...')

df_tenancy_schedule = pd.read_excel(path_tenancy_schedule, sheet_name = 'Report1')
df_tenancy_schedule.columns = df_tenancy_schedule.iloc[1]
df_tenancy_schedule = df_tenancy_schedule[3:]
df_tenancy_schedule = df_tenancy_schedule[ df_tenancy_schedule[' Property '].notnull() ]
df_tenancy_schedule[' Property '] = df_tenancy_schedule[' Property '].apply( lambda x: re.search( r'(.+) - (\d+) (\(\d+\))', x ).groups()[1] )


# In[ ]:


## df_tenancy_schedule


# In[ ]:


df = pd.merge( left = df, right = df_tenancy_schedule[ [' Property ', ' Unit(s) ', ' DBA ', ' Sales Group '] ], on = [' Property ', ' Unit(s) '], how = 'left')
df.drop_duplicates( subset = [' Property ', ' Unit(s) ', ' Lease '], keep = 'last', inplace = True )
df = df.reset_index(drop = True)

df[' DBA Name '] = df[' DBA ']
df.drop(' DBA ', axis = 1, inplace = True)
df[' Tenant Classification '] = df[' Sales Group ']
df.drop(' Sales Group ', axis = 1, inplace = True)

print('Finished.')


# In[ ]:


## df


# In[ ]:


## to update Unit Vacancy
print('Updating Unit Vacancy ...')

df_vacancy = pd.read_excel(path_unit_vacancy, sheet_name = 'Report1')
df_vacancy.columns = df_vacancy.iloc[1]
df_vacancy = df_vacancy[3:]
df_vacancy = df_vacancy[ df_vacancy[' Days '].notnull() ]
df_vacancy = df_vacancy.rename( columns = {' Unit ': ' Unit(s) '} )


# In[ ]:


## df_vacancy


# In[ ]:


df = pd.merge( left = df, right = df_vacancy[ [' Property ', ' Unit(s) ', ' Days '] ], on = [' Property ', ' Unit(s) '], how = 'left')
df.drop_duplicates( subset = [' Property ', ' Unit(s) ', ' Lease '], keep = 'last', inplace = True )
df = df.reset_index(drop = True)

print('Finished.')


# In[ ]:


## df


# In[ ]:


## 
print('Updating Others ...')

## Sorting would disrupt the original structure of dataframe.
## df[' Property '] = df[' Property '].astype(np.int64)
## df.sort_values(by = ' Property ', ascending = True, inplace = True)
## df[' Property '] = df[' Property '].astype(str)

df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' DBA Name '] = ''
df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Legacy/Acquired/New '] = ''
df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Tenant Classification '] = ''
df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Public/Private '] = ''
df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Franchise (Y/N) '] = ''
df.loc[ df[' Lease '].apply(lambda x: 'VACANT' in x), ' Single Tenant Building '] = ''

df[' Count '] = [ x + 1 for x in range(num_row) ]
num_col = df.shape[1]
df = df.reset_index(drop = True)

##
df[' DBA Name '] = df[' DBA Name '].astype(str)
df[' DBA Name '] = np.where( df[' DBA Name '] == 'nan', '', df[' DBA Name '] )
df[' Lease '] = df[' Lease '].astype(str)
df[' Lease '] = np.where( df[' Lease '] == 'nan', '', df[' Lease '] )


# In[ ]:


##
## id_plug = df[' Unit(s) '].apply( lambda x: 'PLUG' in str(x) )
## df.loc[id_plug, 'Suite Size Classification'] = 'Plug'

##
## id_pad1 = df[' Unit(s) '].apply( lambda x: 'PAD' in str(x) )
## id_pad2 = df[' Lease Type '].apply( lambda x: 'Pad' in str(x) )
## id_pad = [ a | b for a, b in zip(id_pad1, id_pad2) ]
## df.loc[id_pad, 'Suite Size Classification'] = 'Pad'

##
## id_roof1 = df[' Unit(s) '].apply( lambda x: 'ROOF' in str(x) )
## df.loc[id_roof1, 'Suite Size Classification'] = 'Roof'
## id_roof2 = df[' Lease '].apply(lambda x: 'VACANT' not in x)
## id_roof = [ a & b for a, b in zip(id_roof1, id_roof2) ]
## df.loc[id_roof, 'Suite Size Classification'] = 'Cell Tower'

##
## str_non_leasable_area = ['HALL', 'OVS', 'RESTROOM', 'GARAGE', 'CA-REST', 'DOCK', 'COOLERS', 'PARKING', 'RISER']
## id_area = df[' Unit(s) '].apply( lambda x: str(x) in str_non_leasable_area )
## df.loc[id_area, 'Suite Size Classification'] = 'Non Leasable Area'

##
## id_signage1 = df[' Lease '].apply(lambda x: 'CBS' in x)
## id_signage2 = df[' Area '].apply(lambda x: x == 0)
## id_signage = [ a & b for a, b in zip(id_signage1, id_signage2) ]
## df.loc[id_signage, 'Suite Size Classification'] = 'Signage'


# In[ ]:


## df


# In[ ]:


## file_origin = r'\\\\EgnyteDrive\\projectphx\\Shared\\ACQUISITIONS DEPT\\Asset Management\\Tenant Database\\Tenant Database 8-6-18 AE.xlsx'
## df_origin = pd.read_excel(file_origin, sheet_name = 'TENANT')
## df_origin.columns = df_origin.iloc[1]
## df_origin = df_origin[2:-2]
## df_origin = df_origin.reset_index(drop = True)
## df_origin = df_origin.rename( columns = {'Lease Name': ' Lease '} )
df_origin['DBA Name'] = df_origin['DBA Name'].astype(str)
df_origin['DBA Name'] = np.where( df_origin['DBA Name'] == 'nan', '', df_origin['DBA Name'] )
df_origin[' Lease '] = df_origin[' Lease '].astype(str)
df_origin[' Lease '] = np.where( df_origin[' Lease '] == 'nan', '', df_origin[' Lease '] )

## to use twice merge to make sure to get index of left only different tenants
df_merge = pd.merge( left = df[ [' Property ', ' Unit(s) ', ' Lease '] ], right = df_origin[ [' Property ', ' Unit(s) ', ' Lease '] ], how = 'outer', indicator = True)
df_merge_left_only = df_merge[ df_merge['_merge'] == 'left_only' ]

df_merge2 = pd.merge( left = df[ [' Property ', ' Unit(s) ', ' Lease '] ], right = df_merge_left_only[ [' Property ', ' Unit(s) ', ' Lease '] ], how = 'outer', indicator = True)
ids_diff = df_merge2[ df_merge2['_merge'] == 'both' ].index.tolist()
## not use ids_true_diff = ids_diff, otherwise no new list, just transfer address
ids_true_diff = ids_diff[:]

for id_ in ids_diff:
    ids_origin = df_origin.index[ df_origin[' Property '] == df.loc[id_, ' Property '] ].tolist()
    
    for id_origin in ids_origin:
        ##if df.loc[id_, 'DBA Name'] != 'nan':
            ##dba = ''.join( re.findall( r'(?i)\b[a-z]+\b', df.loc[id_, 'DBA Name'] ) ).lower()
        ##else:
            ##dba = ''
        ##if df_origin.loc[id_origin, 'DBA Name'] != 'nan':
            ##dba_origin = ''.join( re.findall( r'(?i)\b[a-z]+\b', df_origin.loc[id_origin, 'DBA Name'] ) ).lower()
        ##else:
            ##dba_origin = ''
        ##if df.loc[id_, ' Lease '] != 'nan':
            ##lease = ''.join( re.findall( r'(?i)\b[a-z]+\b', df.loc[id_, ' Lease '] ) ).lower()
        ##else:
            ##lease = ''
        ##if df_origin.loc[id_origin, ' Lease '] != 'nan':
            ##lease_origin = ''.join( re.findall( r'(?i)\b[a-z]+\b', df_origin.loc[id_origin, ' Lease '] ) ).lower()
        ##else:
            ##lease_origin = ''
        dba = ''.join( re.findall( r'(?i)\b[a-z]+\b', df.loc[id_, ' DBA Name '] ) ).lower()
        dba_origin = ''.join( re.findall( r'(?i)\b[a-z]+\b', df_origin.loc[id_origin, 'DBA Name'] ) ).lower()
        lease = ''.join( re.findall( r'(?i)\b[a-z]+\b', df.loc[id_, ' Lease '] ) ).lower()
        lease_origin = ''.join( re.findall( r'(?i)\b[a-z]+\b', df_origin.loc[id_origin, ' Lease '] ) ).lower()
        
        judge_dba = ( (dba in dba_origin) & (dba != '') ) | ( (dba_origin in dba) & (dba_origin != '') )
        judge_lease = ( (lease in lease_origin) & (lease != '') ) | ( (lease_origin in lease) & (lease_origin != '') )
        judge_dba_lease = ( SequenceMatcher(None, dba, dba_origin).ratio() >= 0.7 ) | ( SequenceMatcher(None, lease, lease_origin).ratio() >= 0.7 )
        judge_vacant = ('vacant' not in dba) & ('vacant' not in lease)
        judge_id = id_ in ids_true_diff
        
        if (judge_dba | judge_lease | judge_dba_lease) & judge_vacant & judge_id:
            df.iloc[ id_, df.columns.get_loc(' Legacy/Acquired/New '):df.columns.get_loc(' Single Tenant Building ') + 1 ] = df_origin.iloc[ id_, df_origin.columns.get_loc('Legacy / Acquired / New'):df_origin.columns.get_loc('Single Tenant Building') + 1 ]
            df.iloc[ id_, df.columns.get_loc('Quick Service / Sit Down'):df.columns.get_loc('Co-Tenancy Clause') + 1 ] = df_origin.iloc[ id_, df_origin.columns.get_loc('Quick Service / Sit Down'):df_origin.columns.get_loc('Co-Tenancy Clause') + 1 ]
            ids_true_diff.remove(id_)

print('Finished. \n')


# In[ ]:


## df_merge_left_only
## print(ids_diff)
## print(ids_true_diff)


# In[ ]:


## to seek different tenants
print('Seeking Different Tenants Since Last Update ... ')

df_last = pd.read_excel(path_tenant_database_last, sheet_name = 'TENANT')
df_last.columns = df_last.iloc[9]
df_last = df_last[10:-1]
df_last[' Property '] = df_last[' Property '].apply( lambda x: re.search( r'(\d+) - (.+)', x ).groups()[0] )
df_last = df_last.rename( columns = {'Lease Name': ' Lease '} )
df_last[' Lease '] = df_last[' Lease '].apply(lambda x: 'VACANT' if 'VACANT' in x else x)
df_last = df_last.reset_index(drop = True)


# In[ ]:


## df_last


# In[ ]:


df_last['DBA Name'] = df_last['DBA Name'].astype(str)
df_last['DBA Name'] = np.where( df_last['DBA Name'] == 'nan', '', df_last['DBA Name'] )
df_last[' Lease '] = df_last[' Lease '].astype(str)
df_last[' Lease '] = np.where( df_last[' Lease '] == 'nan', '', df_last[' Lease '] )

## to use twice merge to make sure to get index of left only different tenants
df_merge = pd.merge( left = df[ [' Property ', ' Unit(s) ', ' Lease '] ], right = df_last[ [' Property ', ' Unit(s) ', ' Lease '] ], how = 'outer', indicator = True)
df_merge_left_only = df_merge[ df_merge['_merge'] == 'left_only' ]

df_merge2 = pd.merge( left = df[ [' Property ', ' Unit(s) ', ' Lease '] ], right = df_merge_left_only[ [' Property ', ' Unit(s) ', ' Lease '] ], how = 'outer', indicator = True)
ids_diff = df_merge2[ df_merge2['_merge'] == 'both' ].index.tolist()
## not use ids_true_diff = ids_diff, otherwise no new list, just transfer address
ids_true_diff = ids_diff[:]

for id_ in ids_diff:
    ids_last = df_last.index[ df_last[' Property '] == df.loc[id_, ' Property '] ].tolist()
    
    for id_last in ids_last:
        dba = ''.join( re.findall( r'(?i)\b[a-z]+\b', df.loc[id_, ' DBA Name '] ) ).lower()
        dba_last = ''.join( re.findall( r'(?i)\b[a-z]+\b', df_last.loc[id_last, 'DBA Name'] ) ).lower()
        lease = ''.join( re.findall( r'(?i)\b[a-z]+\b', df.loc[id_, ' Lease '] ) ).lower()
        lease_last = ''.join( re.findall( r'(?i)\b[a-z]+\b', df_last.loc[id_last, ' Lease '] ) ).lower()
        
        judge_dba = ( (dba in dba_last) & (dba != '') ) | ( (dba_last in dba) & (dba_last != '') )
        judge_lease = ( (lease in lease_last) & (lease != '') ) | ( (lease_last in lease) & (lease_last != '') )
        judge_dba_lease = ( SequenceMatcher(None, dba, dba_last).ratio() >= 0.7 ) | ( SequenceMatcher(None, lease, lease_last).ratio() >= 0.7 )
        judge_vacant = ('vacant' not in dba) & ('vacant' not in lease)
        judge_id = id_ in ids_true_diff
        
        if (judge_dba | judge_lease | judge_dba_lease) & judge_vacant & judge_id:
            ids_true_diff.remove(id_)

print('Finished. \n')


# In[ ]:


## ids_true_diff
## df
## print(ids_diff)
## print(ids_true_diff)


# In[ ]:


## to update future tenants
id_future = []
for row in range(num_row):
    if ( df.loc[row, ' Lease '] == 'VACANT' ) & ( df.loc[row, ' Area '] > 0 ):
        for i in range( df_AR_future.shape[0] ):
            if df.loc[row, ' Property '] == df_AR_future.loc[i, ' Property ']:
                id_future.append(row)
                df.loc[row, ' Lease '] = 'VACANT' + df_AR_future.loc[i, ' Lease ']
                df_AR_future.drop( df_AR_future.index[i], inplace = True )
                df_AR_future = df_AR_future.reset_index(drop = True)
                break


# In[ ]:


##
df.columns = df_origin.columns
df[' Property '] = df[' Property '].apply( lambda x: dic_property[x][0] )


# In[ ]:


## df


# In[ ]:


print('Saving to Excel ...')

wb = load_workbook(path_tenant_database_origin)
## wb = load_workbook(r'Z:\Shared\ACQUISITIONS DEPT\Asset Management\Tenant Database\Tenant Database 07-09-2018.xlsx')
sheet = wb['TENANT']

num_row_add = 12

sheet.cell(row = 9, column = 3).value = 'Different Tenant since update {}'.format(date_last)
sheet.cell(row = 2, column = 6).value = str( datetime.now().date() )

##
if df.shape[0] > df_origin.shape[0]:
    sheet.insert_rows( idx = 1000, amount = df.shape[0] - df_origin.shape[0] )
    ## for row in range( 1000, 1000 + df.shape[0] - df_origin.shape[0] ):
        ## for col in range(num_col):
            ## sheet.cell(row = row, column = col + 1).font = copy( sheet.cell(row = row - 1, column = col + 1).font )
            ## sheet.cell(row = row, column = col + 1).number_format = copy( sheet.cell(row = row - 1, column = col + 1).number_format )
            ## sheet.cell(row = row, column = col + 1).alignment = copy( sheet.cell(row = row - 1, column = col + 1).alignment )
elif df.shape[0] < df_origin.shape[0]:
    sheet.delete_rows( idx = 1000, amount = df_origin.shape[0] - df.shape[0] )

##
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Count') + 1 ).value = '= SUBTOTAL(2, A7:A{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Area') + 1 ).value = '= SUBTOTAL(9, N7:N{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Term (Month)') + 1 ).value = '= SUBTOTAL(9, R7:R{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Monthly Rent') + 1 ).value = '= SUBTOTAL(9, S7:S{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Annual Rent') + 1 ).value = '= SUBTOTAL(9, U7:U{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Security Deposit') + 1 ).value = '= SUBTOTAL(9, Y7:Y{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('LOC Amount / Bank Guarantee') + 1 ).value = '= SUBTOTAL(9, Z7:Z{})'.format(num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Current AR Balance') + 1 ).value = '= SUBTOTAL(9, AA7:AA{})'.format(num_row + num_row_add - 1)

letter_col_vacancy_days = get_column_letter( df.columns.get_loc('# of Days Vacant') + 1 )
sum_vacancy_days = '= SUBTOTAL(9, {}7:{}{})'.format(letter_col_vacancy_days, letter_col_vacancy_days, num_row + num_row_add - 1)
sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('# of Days Vacant') + 1 ).value = sum_vacancy_days

sheet.cell( row = num_row + num_row_add, column = df.columns.get_loc('Annual Rent Per Area') + 1 ).value = '= U{}/N{}'.format(num_row + num_row_add, num_row + num_row_add)

## sheet.cell(row = num_row + num_row_add, column = col + 1).border = Border( top = Side(border_style = 'thick') )

## sheet.cell(row = num_row + num_row_add, column = 1).value = num_row
## sheet.cell(row = num_row + num_row_add, column = 14).value = np.nansum( df[' Area '] )
## sheet.cell(row = num_row + num_row_add, column = 19).value = np.nansum( df.iloc[:, 18] )
## sheet.cell(row = num_row + num_row_add, column = 21).value = np.nansum( df.iloc[:, 20] )
## sheet.cell(row = num_row + num_row_add, column = 22).value = np.nansum( df.iloc[:, 20] ) / np.nansum( df[' Area '] )
## sheet.cell(row = num_row + num_row_add, column = 25).value = np.nansum( df[' Security '] )
## sheet.cell(row = num_row + num_row_add, column = 27).value = np.nansum( df['Current'] )

##
for row in range(num_row):
    i = 0
    for col in range(num_col):
        sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fill_type = None)
    if row in id_future:
        for col in range(num_col):
            color = sheet.cell(row = 10, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
    elif ( datetime.now() - df.loc[row, ' Lease From '] ).days <= 30:
        for col in range(num_col):
            color = sheet.cell(row = 3, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
            ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = '32cd32', fill_type = 'solid')
    elif ( ( datetime.now() - df.loc[row, ' Lease From '] ).days > 30 ) & ( ( datetime.now() - df.loc[row, ' Lease From '] ).days <= 90 ):
        for col in range(num_col):
            color = sheet.cell(row = 4, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
            ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = '7cfc00', fill_type = 'solid')
    elif ( ( datetime.now() - df.loc[row, ' Lease From '] ).days > 90 ) & ( ( datetime.now() - df.loc[row, ' Lease From '] ).days <= 120 ):
        for col in range(num_col):
            color = sheet.cell(row = 5, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
            ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = '98fb98', fill_type = 'solid')
    elif ( df.loc[row, ' Lease To '] - datetime.now() ).days <= 30:
        for col in range(num_col):
            color = sheet.cell(row = 6, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
            ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = 'ffd700', fill_type = 'solid')
    elif ( ( df.loc[row, ' Lease To '] - datetime.now() ).days > 30 ) & ( ( df.loc[row, ' Lease To '] - datetime.now() ).days <= 90 ):
        for col in range(num_col):
            color = sheet.cell(row = 7, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
            ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = 'eedd82', fill_type = 'solid')
    elif ( ( df.loc[row, ' Lease To '] - datetime.now() ).days > 90 ) & ( ( df.loc[row, ' Lease To '] - datetime.now() ).days <= 180 ):
        for col in range(num_col):
            color = sheet.cell(row = 8, column = 1).fill.start_color.rgb
            sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
            i += 1
            ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = 'fafad2', fill_type = 'solid')
    ## elif df.loc[row, ' Days '] <= 90:
    ## elif df.loc[row, ' Days '] <= lapse_last.days:
    if row in ids_true_diff:
        color = sheet.cell(row = 9, column = 1).fill.start_color.rgb
        if i == 0:
            for col in range(num_col):
                sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
        else:
            for col in range(num_col)[1::2]:
                sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = color, fill_type = 'solid')
                ## sheet.cell(row = row + num_row_add, column = col + 1).fill = PatternFill(fgColor = 'ffc0cb', fill_type = 'solid')

##
for row in range(num_row - 1):
    if df.loc[row, ' Property '] != df.loc[row + 1, ' Property ']:
        for col in range(num_col): 
            sheet.cell(row = row + num_row_add, column = col + 1).border = Border( bottom = Side(border_style = 'thin') )
    else:
        for col in range(num_col): 
            sheet.cell(row = row + num_row_add, column = col + 1).border = Border( bottom = Side(border_style = None) )
            
##
## df[' Lease From '] = df[' Lease From '].apply( lambda x: x.strftime('%#m/%#d/%Y') )
## df[' Lease From '] = np.where( df[' Lease From '] == datetime.min.strftime('%#m/%#d/%Y'), np.NaN, df[' Lease From '] )
## df[' Lease To '] = df[' Lease To '].apply( lambda x: x.strftime('%#m/%#d/%Y') )
## df[' Lease To '] = np.where( df[' Lease To '] == datetime.max.strftime('%#m/%#d/%Y'), np.NaN, df[' Lease To '] )

df[' Lease From '] = pd.to_datetime( df[' Lease From '], yearfirst = True, errors = 'coerce' ).apply( lambda x: x.date() )
df[' Lease To '] = pd.to_datetime( df[' Lease To '], yearfirst = True, errors = 'coerce' ).apply( lambda x: x.date() )

##
for row in range(num_row):
    for col in range(num_col):
        sheet.cell(row = row + num_row_add, column = col + 1).value = df.iloc[row, col]
        sheet.cell(row = row + num_row_add, column = col + 1).font = Font(name = 'Tahoma', size = 8)
        if col in range(18, 27):
            sheet.cell(row = row + num_row_add, column = col + 1).alignment = Alignment(horizontal = 'right', vertical = 'center', wrap_text = True)
        else:
            sheet.cell(row = row + num_row_add, column = col + 1).alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            
end = datetime.now()
lapse = end - start
print(start, end, lapse)        
        
sheet.cell(row = 3, column = 6).value = 'Update starts at'
sheet.cell(row = 3, column = 7).value = start
sheet.cell(row = 4, column = 6).value = 'Update ends at'
sheet.cell(row = 4, column = 7).value = end
sheet.cell(row = 5, column = 6).value = 'Lapse'
sheet.cell(row = 5, column = 7).value = lapse
                                                                                                       
wb.save( path + 'Tenant Database ' + datetime.now().strftime('%m-%d-%Y') + '.xlsx' )
## str( datetime.now().date() )

print('Finished. \n')


# In[ ]:


end = datetime.now()
lapse = end - start
print(start, end, lapse)


# In[ ]:


## df1 = pd.read_excel('\\\\EgnyteDrive\\projectphx\\Shared\\ACQUISITIONS DEPT\\Asset Management\\Tenant Database\\Tenant Database 08-09-2018.xlsx', sheet_name = 'TENANT')
## df1.columns = df1.iloc[9]
## df1 = df1.iloc[10:-1, 1:]
## df1 = df1.reset_index(drop = True)

## df2 = pd.read_excel(r'C:\Users\acquisitionsdata\Desktop\Tenant Database 08-09-2018.xlsx', sheet_name = 'TENANT')
## df2.columns = df2.iloc[9]
## df2 = df2.iloc[10:-1, 1:]
## df2 = df2.reset_index(drop = True)


# In[ ]:


## df_merge = pd.merge( left = df1.iloc[:, 14:], right = df2.iloc[:, 14:], how = 'outer', indicator = True)
## df_merge_left_only = df_merge[ df_merge['_merge'] == 'left_only' ]
## print(df_merge_left_only)
## df_merge_right_only = df_merge[ df_merge['_merge'] == 'right_only' ]
## print(df_merge_right_only)

